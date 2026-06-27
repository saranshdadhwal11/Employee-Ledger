import os
from datetime import datetime
from functools import wraps
from pathlib import Path

from dotenv import load_dotenv
from flask import (Flask, flash, redirect, render_template, request,
                    send_file, session, url_for)
from openpyxl import Workbook, load_workbook
from werkzeug.security import check_password_hash, generate_password_hash
from pymongo import MongoClient

BASE_DIR = Path(__file__).resolve().parent
env_path = BASE_DIR / ".env"
if not env_path.exists():
    print(f"Warning: .env file not found at {env_path}. Email alerts will not work until .env is created.")
load_dotenv(env_path)

from mailer import send_login_alert

DATA_DIR = BASE_DIR / "data"
USERS_FILE = DATA_DIR / "users.xlsx"
EMPLOYEES_FILE = DATA_DIR / "employees.xlsx"

USER_HEADERS = ["ID", "Username", "Email", "PasswordHash", "CreatedAt"]
EMP_HEADERS = ["ID", "Name", "Department", "Position", "Salary", "Email", "DateJoined"]

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "dev-secret-change-me")

# ----------------------------- MongoDB setup -----------------------------
MONGO_URI = os.environ.get("MONGO_URI", "mongodb://localhost:27017")
MONGO_DB = os.environ.get("MONGO_DB", "ems")
try:
    mongo_client = MongoClient(MONGO_URI, serverSelectionTimeoutMS=2000)
    # trigger server selection to surface errors early
    mongo_client.server_info()
    mongo_db = mongo_client[MONGO_DB]
except Exception:
    mongo_client = None
    mongo_db = None


# ---------------------------------------------------------------- setup ----
def init_excel_files():
    DATA_DIR.mkdir(exist_ok=True)

    if not USERS_FILE.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Users"
        ws.append(USER_HEADERS)
        wb.save(USERS_FILE)

    if not EMPLOYEES_FILE.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Employees"
        ws.append(EMP_HEADERS)
        sample = [
            (1, "Aarav Sharma", "Engineering", "Software Developer", 65000, "aarav.sharma@example.com", "2023-06-12"),
            (2, "Priya Verma", "Human Resources", "HR Executive", 48000, "priya.verma@example.com", "2022-11-03"),
            (3, "Rohan Mehta", "Sales", "Sales Manager", 72000, "rohan.mehta@example.com", "2021-04-20"),
        ]
        for row in sample:
            ws.append(row)
        wb.save(EMPLOYEES_FILE)


# ------------------------------------------------------------ user data ----
def get_users():
    wb = load_workbook(USERS_FILE)
    ws = wb.active
    users = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        users.append({
            "id": row[0], "username": row[1], "email": row[2],
            "password_hash": row[3], "created_at": row[4],
        })
    return users


def find_user_by_username(username):
    for u in get_users():
        if u["username"].lower() == username.lower():
            return u
    return None


def add_user(username, email, password):
    wb = load_workbook(USERS_FILE)
    ws = wb.active
    ids = [r[0].value for r in ws.iter_rows(min_row=2) if r[0].value is not None]
    next_id = max(ids) + 1 if ids else 1
    ws.append([
        next_id, username, email, generate_password_hash(password),
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    ])
    wb.save(USERS_FILE)


# -------------------------------------------------------- employee data ----
def log_employee_operation(operation, emp_id, employee_data):
    """Log employee operation to MongoDB for audit trail."""
    if mongo_db is not None:
        try:
            from flask import session
            username = session.get("user", "unknown")
            mongo_db.employee_operations.insert_one({
                "operation": operation,  # "add", "edit", "delete"
                "emp_id": emp_id,
                "employee_data": employee_data,
                "performed_by": username,
                "timestamp": datetime.utcnow(),
            })
        except Exception:
            pass  # fail silently


def backup_employees_xlsx():
    """Auto-backup employees.xlsx to data/employees_backup.xlsx."""
    try:
        import shutil
        backup_file = DATA_DIR / "employees_backup.xlsx"
        shutil.copy(EMPLOYEES_FILE, backup_file)
    except Exception:
        pass  # fail silently


def get_employees():
    wb = load_workbook(EMPLOYEES_FILE)
    ws = wb.active
    employees = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        employees.append({
            "id": row[0], "name": row[1], "department": row[2],
            "position": row[3], "salary": row[4], "email": row[5],
            "date_joined": row[6],
        })
    return employees


def get_employee(emp_id):
    return next((e for e in get_employees() if int(e["id"]) == int(emp_id)), None)


def add_employee(data):
    wb = load_workbook(EMPLOYEES_FILE)
    ws = wb.active
    ids = [r[0].value for r in ws.iter_rows(min_row=2) if r[0].value is not None]
    next_id = max(ids) + 1 if ids else 1
    ws.append([
        next_id, data["name"], data["department"], data["position"],
        data["salary"], data["email"], data["date_joined"],
    ])
    wb.save(EMPLOYEES_FILE)
    
    # Log to MongoDB and backup Excel
    employee_record = {"id": next_id, **data}
    log_employee_operation("add", next_id, employee_record)
    backup_employees_xlsx()


def update_employee(emp_id, data):
    wb = load_workbook(EMPLOYEES_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if row[0].value is not None and int(row[0].value) == int(emp_id):
            row[1].value, row[2].value, row[3].value = data["name"], data["department"], data["position"]
            row[4].value, row[5].value, row[6].value = data["salary"], data["email"], data["date_joined"]
            break
    wb.save(EMPLOYEES_FILE)
    
    # Log to MongoDB and backup Excel
    employee_record = {"id": emp_id, **data}
    log_employee_operation("edit", emp_id, employee_record)
    backup_employees_xlsx()


def delete_employee(emp_id):
    wb = load_workbook(EMPLOYEES_FILE)
    ws = wb.active
    emp_id = int(emp_id)
    row_to_delete = None
    for row in ws.iter_rows(min_row=2):
        if row[0].value is not None and int(row[0].value) == emp_id:
            row_to_delete = row[0].row
            break

    if row_to_delete is None:
        raise ValueError(f"Employee with ID {emp_id} not found.")

    ws.delete_rows(row_to_delete, 1)
    wb.save(EMPLOYEES_FILE)

    # Log to MongoDB and backup Excel
    log_employee_operation("delete", emp_id, {"id": emp_id})
    backup_employees_xlsx()


# ----------------------------------------------------------------- auth ----
def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if "user" not in session:
            flash("Please log in to continue.", "warning")
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return wrapper


# --------------------------------------------------------------- routes ----
@app.route("/")
def home():
    return redirect(url_for("dashboard") if "user" in session else url_for("login"))


@app.route("/signup", methods=["GET", "POST"])
def signup():
    if request.method == "POST":
        username = request.form["username"].strip()
        email = request.form["email"].strip()
        password = request.form["password"]

        if not username or not email or not password:
            flash("All fields are required.", "danger")
            return redirect(url_for("signup"))
        if find_user_by_username(username):
            flash("That username is already taken.", "danger")
            return redirect(url_for("signup"))

        add_user(username, email, password)
        flash("Account created. Please log in.", "success")
        return redirect(url_for("login"))

    return render_template("signup.html")


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form["username"].strip()
        password = request.form["password"]
        user = find_user_by_username(username)

        if user and check_password_hash(user["password_hash"], password):
            session["user"] = user["username"]
            session["email"] = user["email"]

            # Log the successful login to MongoDB (if available)
            if mongo_db is not None:
                try:
                    mongo_db.logins.update_one(
                        {"username": user["username"]},
                        {
                            "$set": {
                                "email": user["email"],
                                "last_login": datetime.utcnow(),
                                "username": user["username"],
                            },
                            "$inc": {"login_count": 1},
                        },
                        upsert=True,
                    )
                except Exception:
                    # fail silently; login should not be blocked by DB issues
                    pass

            sent, error = send_login_alert(user["email"], user["username"])
            if sent:
                flash(f"Login successful. A confirmation email was sent to {user['email']}.", "stamp")
            else:
                flash(
                    "Login successful, but the notification email could not be sent. "
                    "Please check ems/.env, verify EMAIL_ADDRESS and EMAIL_PASSWORD, "
                    "and ensure you are using a valid Gmail App Password. "
                    f"Error: {error}",
                    "warning",
                )

            return redirect(url_for("dashboard"))

        flash("Invalid username or password.", "danger")

    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    flash("You have been logged out.", "info")
    return redirect(url_for("login"))


@app.route("/dashboard")
@login_required
def dashboard():
    employees = get_employees()
    return render_template("dashboard.html", employees=employees)


@app.route("/employees/add", methods=["GET", "POST"])
@login_required
def add_employee_route():
    if request.method == "POST":
        try:
            salary = int(request.form["salary"])
        except ValueError:
            flash("Salary must be a whole number.", "danger")
            return redirect(url_for("add_employee_route"))

        data = {
            "name": request.form["name"].strip(),
            "department": request.form["department"].strip(),
            "position": request.form["position"].strip(),
            "salary": salary,
            "email": request.form["email"].strip(),
            "date_joined": request.form["date_joined"],
        }
        add_employee(data)
        flash("Employee added.", "success")
        return redirect(url_for("dashboard"))
    return render_template("employee_form.html", employee=None, action="Add")


@app.route("/employees/edit/<int:emp_id>", methods=["GET", "POST"])
@login_required
def edit_employee_route(emp_id):
    employee = get_employee(emp_id)
    if not employee:
        flash("Employee not found.", "danger")
        return redirect(url_for("dashboard"))

    if request.method == "POST":
        try:
            salary = int(request.form["salary"])
        except ValueError:
            flash("Salary must be a whole number.", "danger")
            return redirect(url_for("edit_employee_route", emp_id=emp_id))

        data = {
            "name": request.form["name"].strip(),
            "department": request.form["department"].strip(),
            "position": request.form["position"].strip(),
            "salary": salary,
            "email": request.form["email"].strip(),
            "date_joined": request.form["date_joined"],
        }
        update_employee(emp_id, data)
        flash("Employee updated.", "success")
        return redirect(url_for("dashboard"))

    return render_template("employee_form.html", employee=employee, action="Edit")


@app.route("/employees/delete/<int:emp_id>", methods=["POST"])
@login_required
def delete_employee_route(emp_id):
    delete_employee(emp_id)
    flash("Employee deleted.", "info")
    return redirect(url_for("dashboard"))


@app.route("/employees/export")
@login_required
def export_employees():
    return send_file(EMPLOYEES_FILE, as_attachment=True, download_name="employees.xlsx")


init_excel_files()

if __name__ == "__main__":
    app.run(debug=True)
